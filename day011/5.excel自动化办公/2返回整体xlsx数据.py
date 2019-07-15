#  xlsx xls
#openpyxl   -> xlsx

from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    dic = {}
    #打开文件
    file = load_workbook(filename=path)
    #表格的名称
    # print(file.get_sheet_names())
    sheets = file.get_sheet_names()

    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        #一张表的所有数据
        sheetInfo = []
        for lineNum in range(1,sheet.max_row+1):
            lineList=[]
            for columnNum in range(1,sheet.max_column+1):
                value = sheet.cell(row=lineNum,column=columnNum).value
                lineList.append(value)  #
            sheetInfo.append(lineList)  #
        #将一张表的数据存到字典
        dic[sheetName] = sheetInfo
    return dic

path = r"G:\python_learn\pycharm_learn\pythonlearn\day011\5.excel自动化办公\tracy1.xlsx"

dic = readXlsxFile(path)
print(dic)
















