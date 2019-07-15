#  xlsx xls
#openpyxl   -> xlsx

from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    #打开文件
    file = load_workbook(filename=path)
    #表格的名称
    # print(file.get_sheet_names())
    sheets = file.get_sheet_names()

    #拿出一个表格
    sheet = file.get_sheet_by_name(sheets[0])
    #最大行数
    # print(sheet.max_row)
    #最大列数
    # print(sheet.max_colum)
    #表名
    # print(sheet.title)

    #读取一张表的数据

    for lineNum in range(1,sheet.max_row+1):
        # print(lineNum)
        lineList=[]
        for columnNum in range(1,sheet.max_column+1):
            #拿数据
            value = sheet.cell(row=lineNum,column=columnNum).value
            lineList.append(value)
        print(lineList)

path = r"G:\python_learn\pycharm_learn\pythonlearn\day011\5.excel自动化办公\tracy1.xlsx"
'''
#它不能处理xls文件
'''
readXlsxFile(path)





#一行为一个列表,放到一个大列表
'''
{"安利播发":[[],[],,,,[]],"安利播发1":[[],[],,,,[]]}

'''